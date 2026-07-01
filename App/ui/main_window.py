import os
import re
import cv2
import sqlite3
from datetime import datetime
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QPushButton, QFileDialog, QLabel, QScrollArea,
    QHBoxLayout, QGridLayout, QMessageBox, QFrame, QProgressBar, QSizePolicy,
    QListWidget, QListWidgetItem, QLineEdit, QStyle, QComboBox
)
from PyQt5.QtWebEngineWidgets import QWebEngineView
from PyQt5.QtWebChannel import QWebChannel
import json
import html
from PyQt5.QtGui import QPixmap, QColor, QTextDocument
from PyQt5.QtCore import Qt, QUrl, QTimer, QObject, pyqtSlot
from PyQt5.QtMultimedia import QMediaPlayer, QMediaContent
from PyQt5.QtMultimediaWidgets import QVideoWidget
from PyQt5.QtPrintSupport import QPrinter

from config import LOGO_PATH, IMAGE_EXTS, VIDEO_EXTS
from utils.image_utils import cv2_to_qpixmap, clear_layout
from ui.widgets import create_card, make_chip, status_color, damage_level_color

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) if os.path.basename(os.path.dirname(os.path.abspath(__file__))) == "ui" else os.path.dirname(os.path.abspath(__file__))
DATABASE_PATH = os.path.join(BASE_DIR, "database", "potholes.db")


REPORT_STATUS_INFO = {
    "pending": {
        "label": "Chờ duyệt",
        "text_color": "#1D4ED8",
        "bg_color": "#DBEAFE",
        "border_color": "#93C5FD",
        "marker_color": "#2563EB",
    },
    "approved": {
        "label": "Đã duyệt",
        "text_color": "#15803D",
        "bg_color": "#DCFCE7",
        "border_color": "#86EFAC",
        "marker_color": "#16A34A",
    },
    "processing": {
        "label": "Đang xử lý",
        "text_color": "#B45309",
        "bg_color": "#FEF3C7",
        "border_color": "#FCD34D",
        "marker_color": "#F59E0B",
    },
    "resolved": {
        "label": "Đã xử lý",
        "text_color": "#15803D",
        "bg_color": "#DCFCE7",
        "border_color": "#86EFAC",
        "marker_color": "#16A34A",
    },
    "need_more": {
        "label": "Cần bổ sung",
        "text_color": "#B45309",
        "bg_color": "#FEF3C7",
        "border_color": "#FCD34D",
        "marker_color": "#F59E0B",
    },
    "invalid": {
        "label": "Không hợp lệ",
        "text_color": "#B91C1C",
        "bg_color": "#FEE2E2",
        "border_color": "#FCA5A5",
        "marker_color": "#DC2626",
    },
}

REPORT_STATUS_ORDER = [
    "pending",
    "need_more",
    "approved",
    "processing",
    "resolved",
    "invalid",
]

REPORT_STATUS_TRANSITIONS = {
    "pending": ["approved", "need_more", "invalid"],
    "need_more": ["approved", "invalid"],
    "approved": ["processing"],
    "processing": ["resolved"],
    "resolved": [],
    "invalid": [],
}
try:
    from app_settings import MAP_HTML_PATH
except Exception:
    MAP_HTML_PATH = os.path.join(BASE_DIR, "map.html")
    
    
# MAIN APP
class AdminMapBridge(QObject):
    def __init__(self, parent=None):
        super().__init__(parent)

    @pyqtSlot(str)
    def selectReport(self, report_id):
        parent = self.parent()
        if parent is not None and hasattr(parent, "select_report_from_map"):
            parent.select_report_from_map(report_id)
        
class YOLOApp(QWidget):
    def __init__(self):
        super().__init__()
        self.selected_files = []
        self.selected_files = []
        self.current_report_id = None
        self.current_report_status = None
        self.report_filter = "all"
        self.report_items = {}
        self.admin_report_window = None
        self.video_workers = []
        self.video_progress_bars = {}
        self.video_status_labels = {}
        self.video_cards = {}
        self.video_players = []
        self.input_video_players = {}
        self.input_video_caps = {}
        self.input_video_timers = {}
        self.input_video_labels = {}
        self.input_preview_player = None
        self.input_preview_video = None
        self.input_video_cap = None
        self.input_video_timer = QTimer(self)
        self.input_video_timer.timeout.connect(self.update_input_video_frame)
        
        self.selected_report_data = None
        self.report_detail_window = None
        
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Pothole Detection - UTC2")
        self.setGeometry(70, 30, 1520, 950)

        self.setStyleSheet("""
            QWidget {
                background-color: #F7F7F7;
                color: #111827;
                font-family: "Segoe UI";
            }
            QLabel {
                background: transparent;
            }
            QPushButton {
                background-color: #F97316;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 16px;
                font-size: 13px;
                font-weight: 700;
            }
            QPushButton:hover {
                background-color: #EA580C;
            }
            QPushButton:pressed {
                background-color: #C2410C;
            }
            QScrollArea {
                border: none;
                background: transparent;
            }
            QFrame#headerBar {
                background: white;
                border: 1px solid #D9D9D9;
                border-radius: 4px;
            }
            QFrame#headerTopBlue {
                background: #0E74B8;
                border: none;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
            }
            QFrame#headerWhite {
                background: white;
                border: none;
            }
            QFrame#panelBox {
                background: white;
                border: 1px solid #D9D9D9;
                border-radius: 4px;
            }
            QFrame#panelTitle {
                background: white;
                border: none;
            }
            QFrame#resultCard {
                background: #FCFCFC;
                border: 1px solid #E2E8F0;
                border-radius: 10px;
            }
            QProgressBar {
                background-color: #E5E7EB;
                border: none;
                border-radius: 8px;
                text-align: center;
                min-height: 18px;
                font-size: 11px;
                font-weight: 700;
                color: #111827;
            }
            QProgressBar::chunk {
                background-color: #16A34A;
                border-radius: 8px;
            }
        """)

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(4)

        # ================= HEADER =================
        header = QFrame()
        header.setObjectName("headerBar")
        header_layout = QVBoxLayout(header)
        header_layout.setContentsMargins(0, 0, 0, 6)
        header_layout.setSpacing(0)

        # ===== KHỐI XANH =====
        blue_bg = QFrame()
        blue_bg.setStyleSheet("""
            QFrame {
                background: #1F6FA5;
                border-radius: 6px;
            }
        """)
        blue_layout = QHBoxLayout(blue_bg)
        blue_layout.setContentsMargins(20, 7, 20, 7)
        blue_layout.setSpacing(16)

        # ===== LOGO (TRONG KHỐI XANH) =====
        logo_box = QFrame()
        logo_box.setFixedSize(110, 110)
        logo_box.setStyleSheet("""
            QFrame {
                background: white;
                border-radius: 18px;
            }
        """)

        logo_layout = QVBoxLayout(logo_box)
        logo_layout.setContentsMargins(8, 8, 8, 8)

        self.logo_label = QLabel()
        self.logo_label.setAlignment(Qt.AlignCenter)

        if os.path.exists(LOGO_PATH):
            pix = QPixmap(LOGO_PATH).scaled(
                90, 90, Qt.KeepAspectRatio, Qt.SmoothTransformation
            )
            self.logo_label.setPixmap(pix)
        else:
            self.logo_label.setText("LOGO")

        logo_layout.addWidget(self.logo_label)
        blue_layout.addWidget(logo_box, 0, Qt.AlignTop)

        # ===== TEXT TRONG KHỐI XANH =====
        text_layout = QVBoxLayout()
        text_layout.setSpacing(4)

        vi_title = QLabel("TRƯỜNG ĐẠI HỌC GIAO THÔNG VẬN TẢI PHÂN HIỆU TẠI THÀNH PHỐ HỒ CHÍ MINH")
        vi_title.setStyleSheet("""
            QLabel {
                color: white;
                font-size: 28px;
                font-weight: 900;
            }
        """)

        en_title = QLabel("UNIVERSITY OF TRANSPORT AND COMMUNICATIONS - CAMPUS IN HO CHI MINH CITY")
        en_title.setStyleSheet("""
            QLabel {
                color: #FFD54A;
                font-size: 16px;
                font-weight: 700;
            }
        """)

        # line đỏ
        red_line = QFrame()
        red_line.setFixedHeight(2)
        red_line.setStyleSheet("background:#D62828;")

        text_layout.addWidget(vi_title)
        text_layout.addWidget(en_title)
        text_layout.addWidget(red_line)

        blue_layout.addLayout(text_layout, 1)

        header_layout.addWidget(blue_bg)

        # ===== DEMO TITLE (NẰM SÁT HEADER) =====
        demo = QLabel("HỆ THỐNG PHÁT HIỆN Ổ GÀ")
        demo.setAlignment(Qt.AlignCenter)
        demo.setStyleSheet("""
            QLabel {
                font-size: 26px;
                font-weight: 900;
                color: #1F2937;
                padding-top: 8px;
                padding-bottom: 2px;
            }
        """)

        header_layout.addWidget(demo)

        main_layout.addWidget(header)        

        # ================= MAIN CONTENT: REPORT LIST + MAP =================
        content_layout = QHBoxLayout()
        content_layout.setSpacing(10)

        # =====================================================
        # LEFT PANEL: DANH SÁCH BÁO CÁO NGƯỜI DÂN
        # =====================================================
        left_panel = QVBoxLayout()
        left_panel.setSpacing(8)

        report_title = QLabel("DANH SÁCH BÁO CÁO NGƯỜI DÂN")
        report_title.setStyleSheet("""
            QLabel {
                font-size: 15px;
                font-weight: 900;
                color: #1F2937;
                padding-left: 2px;
            }
        """)
        left_panel.addWidget(report_title)

        report_box = QFrame()
        report_box.setObjectName("panelBox")

        report_box_layout = QVBoxLayout(report_box)
        report_box_layout.setContentsMargins(12, 12, 12, 12)
        report_box_layout.setSpacing(10)

        # ===== Bộ lọc trạng thái báo cáo =====
        filter_caption = QLabel("Lọc theo trạng thái")
        filter_caption.setStyleSheet("""
            QLabel {
                color: #334155;
                font-size: 12px;
                font-weight: 900;
                padding-left: 2px;
            }
        """)
        report_box_layout.addWidget(filter_caption)

        report_filter_row_1 = QHBoxLayout()
        report_filter_row_1.setSpacing(8)
        report_filter_row_2 = QHBoxLayout()
        report_filter_row_2.setSpacing(8)

        self.btn_report_all = QPushButton("Tất cả")
        self.btn_report_pending = QPushButton("Chờ duyệt")
        self.btn_report_approved = QPushButton("Đã duyệt")
        self.btn_report_processing = QPushButton("Đang xử lý")
        self.btn_report_resolved = QPushButton("Đã xử lý")
        self.btn_report_need_more = QPushButton("Cần bổ sung")
        self.btn_report_invalid = QPushButton("Không hợp lệ")

        self.status_filter_buttons = {
            "all": self.btn_report_all,
            "pending": self.btn_report_pending,
            "approved": self.btn_report_approved,
            "processing": self.btn_report_processing,
            "resolved": self.btn_report_resolved,
            "need_more": self.btn_report_need_more,
            "invalid": self.btn_report_invalid,
        }

        for filter_value, button in self.status_filter_buttons.items():
            button.clicked.connect(lambda checked=False, value=filter_value: self.set_report_filter(value))
            button.setMinimumHeight(34)

        report_filter_row_1.addWidget(self.btn_report_all)
        report_filter_row_1.addWidget(self.btn_report_pending)
        report_filter_row_1.addWidget(self.btn_report_approved)
        report_filter_row_1.addWidget(self.btn_report_processing)

        report_filter_row_2.addWidget(self.btn_report_resolved)
        report_filter_row_2.addWidget(self.btn_report_need_more)
        report_filter_row_2.addWidget(self.btn_report_invalid)
        report_filter_row_2.addStretch(1)

        report_box_layout.addLayout(report_filter_row_1)
        report_box_layout.addLayout(report_filter_row_2)

        # ===== Xuất thống kê báo cáo =====
        export_frame = QFrame()
        export_frame.setStyleSheet("""
            QFrame {
                background: #F8FAFC;
                border: 1px solid #E2E8F0;
                border-radius: 8px;
            }
        """)
        export_layout = QHBoxLayout(export_frame)
        export_layout.setContentsMargins(8, 8, 8, 8)
        export_layout.setSpacing(8)

        export_label = QLabel("Xuất thống kê")
        export_label.setStyleSheet("""
            QLabel {
                color: #334155;
                font-size: 12px;
                font-weight: 900;
            }
        """)

        self.cmb_export_status = QComboBox()
        self.cmb_export_status.addItem("Theo bộ lọc hiện tại", "current")
        self.cmb_export_status.addItem("Tất cả", "all")
        for status_value in REPORT_STATUS_ORDER:
            self.cmb_export_status.addItem(REPORT_STATUS_INFO[status_value]["label"], status_value)
        self.cmb_export_status.setStyleSheet("""
            QComboBox {
                background: #FFFFFF;
                border: 1px solid #CBD5E1;
                border-radius: 8px;
                padding: 7px 10px;
                font-size: 12px;
                font-weight: 700;
                min-height: 24px;
            }
        """)

        self.btn_export_excel = QPushButton("Xuất Excel")
        self.btn_export_excel.setToolTip("Xuất danh sách báo cáo theo trạng thái đang chọn ra file Excel")
        self.btn_export_excel.clicked.connect(self.export_reports_to_excel)
        self.btn_export_excel.setStyleSheet("""
            QPushButton {
                background-color: #16A34A;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 8px 12px;
                font-size: 12px;
                font-weight: 900;
            }
            QPushButton:hover { background-color: #15803D; }
        """)

        self.btn_export_pdf = QPushButton("Xuất PDF")
        self.btn_export_pdf.setToolTip("Xuất danh sách báo cáo theo trạng thái đang chọn ra file PDF")
        self.btn_export_pdf.clicked.connect(self.export_reports_to_pdf)
        self.btn_export_pdf.setStyleSheet("""
            QPushButton {
                background-color: #DC2626;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 8px 12px;
                font-size: 12px;
                font-weight: 900;
            }
            QPushButton:hover { background-color: #B91C1C; }
        """)

        export_layout.addWidget(export_label)
        export_layout.addWidget(self.cmb_export_status, 1)
        export_layout.addWidget(self.btn_export_excel)
        export_layout.addWidget(self.btn_export_pdf)

        report_box_layout.addWidget(export_frame)

        # ===== List báo cáo =====
        self.report_list_widget = QListWidget()
        self.report_list_widget.itemClicked.connect(self.on_citizen_report_clicked)
        self.report_list_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.report_list_widget.setStyleSheet("""
            QListWidget {
                background: white;
                border: 1px solid #E5E7EB;
                border-radius: 8px;
                padding: 6px;
                font-size: 13px;
                color: #111827;
            }
            QListWidget::item {
                padding: 9px;
                border-bottom: 1px solid #F1F5F9;
            }
            QListWidget::item:selected {
                background: #DBEAFE;
                color: #111827;
                border-radius: 6px;
            }
        """)

        report_box_layout.addWidget(self.report_list_widget, 1)

        # ===== Chỉ giữ nút mở trang chi tiết
        # Phần hiển thị thông tin báo cáo và cập nhật trạng thái đã được chuyển sang cửa sổ chi tiết.

        self.btn_view_report_detail = QPushButton("Xem / cập nhật chi tiết báo cáo")
        self.btn_view_report_detail.setEnabled(False)
        self.btn_view_report_detail.clicked.connect(self.open_report_detail_window)
        self.btn_view_report_detail.setStyleSheet("""
            QPushButton {
                background-color: #2563EB;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 16px;
                font-size: 13px;
                font-weight: 800;
            }
            QPushButton:hover {
                background-color: #1D4ED8;
            }
            QPushButton:disabled {
                background-color: #CBD5E1;
                color: #64748B;
            }
        """)

        report_box_layout.addWidget(self.btn_view_report_detail)

        # ===== Log nhỏ bên dưới =====
        self.log_label = QLabel("Logs")
        self.log_label.setStyleSheet("""
            QLabel {
                color: #6B7280;
                font-size: 11px;
                font-weight: 700;
                border-top: 1px solid #E5E7EB;
                padding-top: 6px;
            }
        """)
        report_box_layout.addWidget(self.log_label)

        self.log_content = QLabel("Chưa có log.")
        self.log_content.setWordWrap(True)
        self.log_content.setMinimumHeight(55)
        self.log_content.setStyleSheet("""
            QLabel {
                border: 1px solid #E5E7EB;
                background: #FAFAFA;
                color: #4B5563;
                font-size: 11px;
                padding: 8px;
            }
        """)
        report_box_layout.addWidget(self.log_content)

        left_panel.addWidget(report_box, 1)

        # RIGHT PANEL: MAP
        right_panel = QVBoxLayout()
        right_panel.setSpacing(8)

        map_title = QLabel("BẢN ĐỒ VỊ TRÍ Ổ GÀ")
        map_title.setStyleSheet("""
            QLabel {
                font-size: 15px;
                font-weight: 900;
                color: #1F2937;
                padding-left: 2px;
            }
        """)
        right_panel.addWidget(map_title)

        map_box = QFrame()
        map_box.setObjectName("panelBox")

        map_box_layout = QVBoxLayout(map_box)
        map_box_layout.setContentsMargins(10, 10, 10, 10)
        map_box_layout.setSpacing(10)

        # ===== Hàng tìm kiếm =====
        map_search_row = QHBoxLayout()
        map_search_row.setSpacing(8)

        self.txt_search = QLineEdit()
        self.txt_search.setPlaceholderText("Nhập tên đường hoặc địa chỉ")
        self.txt_search.setStyleSheet("""
            QLineEdit {
                background: white;
                border: 1px solid #CBD5E1;
                border-radius: 8px;
                padding: 9px 10px;
                font-size: 13px;
            }
            QLineEdit:focus {
                border: 2px solid #2563EB;
            }
        """)

        self.btn_search = QPushButton()
        self.btn_search.setToolTip("Tìm kiếm địa chỉ")
        self.btn_search.setIcon(self.style().standardIcon(QStyle.SP_FileDialogContentsView))
        self.btn_search.setFixedWidth(50)
        self.btn_search.clicked.connect(self.search_location_by_button)
        self.btn_search.setStyleSheet("""
            QPushButton {
                background-color: #FFFFFF;
                color: #111827;
                border: 1px solid #CBD5E1;
                border-radius: 8px;
                padding: 8px;
            }
            QPushButton:hover {
                background-color: #F8FAFC;
            }
        """)

        map_search_row.addWidget(self.txt_search, 1)
        map_search_row.addWidget(self.btn_search)

        map_box_layout.addLayout(map_search_row)

        # ===== Web map =====
        self.web_view = QWebEngineView()
        self.web_view.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        # Quan trọng: gắn QWebChannel trước khi load map.html
        self.channel = QWebChannel()
        self.bridge = AdminMapBridge(self)
        self.channel.registerObject("backend", self.bridge)
        self.web_view.page().setWebChannel(self.channel)

        self.web_view.loadFinished.connect(self.on_map_loaded)

        if not os.path.exists(MAP_HTML_PATH):
            QMessageBox.critical(
                self,
                "Thiếu file map.html",
                f"Không tìm thấy file bản đồ:\n{MAP_HTML_PATH}"
            )
        else:
            self.web_view.load(QUrl.fromLocalFile(MAP_HTML_PATH))

        map_box_layout.addWidget(self.web_view, 1)

        right_panel.addWidget(map_box, 1)

        # =====================================================
        # ADD TO MAIN CONTENT
        # =====================================================
        content_layout.addLayout(left_panel, 1)
        content_layout.addLayout(right_panel, 1)

        main_layout.addLayout(content_layout, 1)

        self.map_loaded = False
        self.report_data_list = []
        self.update_report_filter_button_styles()
        self.load_citizen_reports()

    # =========================
    # UI HELPERS
    # =========================
    def set_log(self, text):
        self.log_content.setText(text)

    def normalize_report_status(self, status):
        raw = str(status or "pending").strip().lower()
        raw_no_space = raw.replace(" ", "_").replace("-", "_")

        aliases = {
            "": "pending",
            "none": "pending",
            "null": "pending",
            "pending": "pending",
            "waiting": "pending",
            "draft": "pending",
            "chờ_duyệt": "pending",
            "cho_duyet": "pending",
            "chua_duyet": "pending",
            "chưa_duyệt": "pending",
            "approved": "approved",
            "da_duyet": "approved",
            "đã_duyệt": "approved",
            "processing": "processing",
            "in_progress": "processing",
            "dang_xu_ly": "processing",
            "đang_xử_lý": "processing",
            "resolved": "resolved",
            "done": "resolved",
            "handled": "resolved",
            "da_xu_ly": "resolved",
            "đã_xử_lý": "resolved",
            "need_more": "need_more",
            "need_supplement": "need_more",
            "supplement_required": "need_more",
            "can_bo_sung": "need_more",
            "cần_bổ_sung": "need_more",
            "invalid": "invalid",
            "rejected": "invalid",
            "khong_hop_le": "invalid",
            "không_hợp_lệ": "invalid",
        }

        if raw_no_space in aliases:
            return aliases[raw_no_space]
        if raw in REPORT_STATUS_INFO:
            return raw
        return "pending"

    def get_report_status_info(self, status):
        return REPORT_STATUS_INFO.get(
            self.normalize_report_status(status),
            REPORT_STATUS_INFO["pending"]
        )

    def get_report_status_label(self, status):
        return self.get_report_status_info(status)["label"]

    def get_report_status_marker_color(self, status):
        return self.get_report_status_info(status)["marker_color"]

    def get_report_status_badge_html(self, status):
        info = self.get_report_status_info(status)
        return (
            f"<span style='background:{info['bg_color']};color:{info['text_color']};"
            f"border:1px solid {info['border_color']};padding:2px 7px;"
            f"border-radius:999px;font-weight:800'>{html.escape(info['label'])}</span>"
        )

    def get_allowed_next_statuses(self, current_status):
        current_status = self.normalize_report_status(current_status)
        return REPORT_STATUS_TRANSITIONS.get(current_status, [])

    def get_status_button_style(self, status_value, enabled=True):
        info = REPORT_STATUS_INFO[status_value]
        if not enabled:
            return """
                QPushButton {
                    background-color: #E5E7EB;
                    color: #94A3B8;
                    border: 1px solid #CBD5E1;
                    border-radius: 8px;
                    padding: 8px 10px;
                    font-size: 12px;
                    font-weight: 800;
                }
            """

        return f"""
            QPushButton {{
                background-color: {info['bg_color']};
                color: {info['text_color']};
                border: 1px solid {info['border_color']};
                border-radius: 8px;
                padding: 8px 10px;
                font-size: 12px;
                font-weight: 900;
            }}
            QPushButton:hover {{
                background-color: #FFFFFF;
                border: 1px solid {info['marker_color']};
            }}
            QPushButton:pressed {{
                background-color: {info['bg_color']};
            }}
        """

    def update_status_action_buttons(self):
        buttons = getattr(self, "status_action_buttons", {})
        if not buttons:
            return

        if self.current_report_id is None or self.current_report_status is None:
            for status_value, button in buttons.items():
                button.setEnabled(False)
                button.setStyleSheet(self.get_status_button_style(status_value, enabled=False))
            return

        current_status = self.normalize_report_status(self.current_report_status)
        allowed_next_statuses = set(self.get_allowed_next_statuses(current_status))

        for status_value, button in buttons.items():
            enabled = status_value in allowed_next_statuses
            button.setEnabled(enabled)
            button.setStyleSheet(self.get_status_button_style(status_value, enabled=enabled))

        current_label = self.get_report_status_label(current_status)
        if current_status in ("resolved", "invalid"):
            self.approve_report_button.setToolTip(f"Báo cáo đã ở trạng thái cuối: {current_label}")
        else:
            self.approve_report_button.setToolTip("Chỉ bật khi trạng thái hiện tại cho phép chuyển sang Đã duyệt.")

    def render_report_detail_text(self, report):
        status = self.normalize_report_status(report.get("status"))
        status_text = self.get_report_status_label(status)
        source_name = report.get("source_name") or self.extract_source_from_report_address(
            report.get("address"),
            report.get("image_path")
        )
        video_time = report.get("video_time") or self.extract_video_time_from_report_address(report.get("address"))
        time_line = f"\nThời điểm video: {video_time}" if self.is_video_source(source_name) and video_time else ""

        return (
            f"Mã báo cáo: #{report['id']}\n"
            f"Trạng thái: {status_text}\n"
            f"Tuyến đường: {report.get('road_name') or self.clean_report_road_name(report.get('address'))}\n"
            f"Nguồn: {source_name}"
            f"{time_line}\n"
            f"Tọa độ: {report['latitude']}, {report['longitude']}\n"
            f"Số ảnh: {report['image_count']}\n"
            f"Thời gian gửi: {report['created_at']}"
        )

    def update_current_report_status(self, new_status):
        if self.current_report_id is None or not self.selected_report_data:
            QMessageBox.warning(
                self,
                "Chưa chọn báo cáo",
                "Vui lòng chọn một báo cáo người dân trước khi cập nhật trạng thái."
            )
            return

        current_status = self.normalize_report_status(self.current_report_status)
        new_status = self.normalize_report_status(new_status)
        allowed_next_statuses = self.get_allowed_next_statuses(current_status)

        if new_status not in allowed_next_statuses:
            QMessageBox.warning(
                self,
                "Không đúng luồng trạng thái",
                f"Không thể chuyển từ '{self.get_report_status_label(current_status)}' "
                f"sang '{self.get_report_status_label(new_status)}'."
            )
            return

        try:
            conn = sqlite3.connect(DATABASE_PATH)
            cursor = conn.cursor()

            cursor.execute("""
                UPDATE pothole_reports
                SET status = ?
                WHERE id = ?
            """, (new_status, self.current_report_id))

            conn.commit()
            conn.close()

            self.current_report_status = new_status
            self.selected_report_data["status"] = new_status

            status_text = self.get_report_status_label(new_status)
            self.set_log(
                f"Đã chuyển báo cáo #{self.current_report_id} sang trạng thái: {status_text}."
            )

            # Nếu đang lọc theo trạng thái cũ thì chuyển về Tất cả để báo cáo vừa cập nhật không biến mất khỏi danh sách.
            if self.report_filter not in ("all", new_status):
                self.report_filter = "all"
                self.update_report_filter_button_styles()

            selected_id = self.current_report_id
            self.load_citizen_reports()
            self.select_report_from_map(selected_id)

            QMessageBox.information(
                self,
                "Cập nhật trạng thái thành công",
                f"Báo cáo #{selected_id} đã chuyển sang trạng thái: {status_text}."
            )

        except Exception as e:
            QMessageBox.critical(
                self,
                "Lỗi cập nhật trạng thái",
                f"Không thể cập nhật trạng thái báo cáo.\n\nChi tiết: {e}"
            )

    def make_section_title(self, text):
        lbl = QLabel(text)
        lbl.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: 800;
                color: #111827;
            }
        """)
        return lbl

    def make_subtitle(self, text, color="#475569"):
        lbl = QLabel(text)
        lbl.setWordWrap(True)
        lbl.setStyleSheet(f"""
            QLabel {{
                color: {color};
                font-size: 12px;
                font-weight: 600;
            }}
        """)
        return lbl

    def chip_row(self, widgets):
        row = QHBoxLayout()
        row.setSpacing(8)
        for w in widgets:
            row.addWidget(w)
        row.addStretch()
        return row
    
    def get_media_source_name(self, path_or_name):
        if not path_or_name:
            return "--"
        return os.path.basename(str(path_or_name).strip()) or "--"

    def is_video_source(self, source_name):
        return os.path.splitext(source_name or "")[1].lower() in VIDEO_EXTS

    def clean_report_road_name(self, address):
        text = str(address or "--").strip()

        if text.startswith("Tuyến đường:"):
            text = text[len("Tuyến đường:"):].strip()

        for token in ("Nguồn:", "Thời điểm video:"):
            if token in text:
                text = text.split(token, 1)[0].strip()

        return text or "--"

    def extract_source_from_report_address(self, address, fallback_path=""):
        text = str(address or "")
        source_name = ""

        if "Nguồn:" in text:
            after_source = text.split("Nguồn:", 1)[1].strip()
            if after_source.lower().startswith("video "):
                after_source = after_source[6:].strip()
            if "Thời điểm video:" in after_source:
                after_source = after_source.split("Thời điểm video:", 1)[0].strip()
            source_name = after_source.strip()

        if not source_name and fallback_path:
            source_name = self.get_media_source_name(fallback_path)

        return self.get_media_source_name(source_name)

    def extract_video_time_from_report_address(self, address, default_time=""):
        text = str(address or "")
        if "Thời điểm video:" not in text:
            return default_time
        return text.split("Thời điểm video:", 1)[1].strip().split()[0]

    def build_admin_report_popup_html(self, report):
        status = self.normalize_report_status(report.get("status"))
        status_badge = self.get_report_status_badge_html(status)

        road_name = report.get("road_name") or self.clean_report_road_name(report.get("address"))
        source_name = report.get("source_name") or self.extract_source_from_report_address(
            report.get("address"),
            report.get("image_path") or report.get("image_name") or ""
        )
        video_time = report.get("video_time") or self.extract_video_time_from_report_address(report.get("address"))
        is_video = self.is_video_source(source_name)

        rows = [
            f"<div style='font-weight:900;color:#2563EB;margin-bottom:7px;font-size:15px'>Báo cáo #{html.escape(str(report.get('id', '--')))}</div>",
            f"<div><b>Trạng thái:</b> {status_badge}</div>",
            f"<div><b>Tuyến đường:</b> {html.escape(str(road_name))}</div>",
            f"<div><b>Nguồn:</b> {html.escape(str(source_name))}</div>",
        ]

        if is_video and video_time:
            rows.append(f"<div><b>Thời điểm video:</b> {html.escape(str(video_time))}</div>")

        rows.append(f"<div><b>Thời gian gửi:</b> {html.escape(str(report.get('created_at') or '--'))}</div>")
        rows.append(f"<div><b>Số file:</b> {int(report.get('image_count') or 0)}</div>")

        lat = report.get("latitude")
        lng = report.get("longitude")
        if lat is not None and lng is not None:
            rows.append(f"<div><b>Tọa độ:</b> {float(lat):.6f}, {float(lng):.6f}</div>")

        return "<div style='min-width:270px;line-height:1.45;font-size:13px;color:#111827'>" + "".join(rows) + "</div>"

    def make_recommendation_box(self, text, damage_level):
        if damage_level in ["High", "Severe"]:
            bg = "#FFF1F2"
            border = "#FB7185"
            color = "#BE123C"
        elif damage_level == "Medium":
            bg = "#FFFBEB"
            border = "#FBBF24"
            color = "#B45309"
        else:
            bg = "#F0FDF4"
            border = "#22C55E"
            color = "#15803D"

        box = QFrame()
        box.setStyleSheet(f"""
            QFrame {{
                background-color: {bg};
                border: 1px solid {border};
                border-radius: 10px;
            }}
        """)

        layout = QHBoxLayout(box)
        layout.setContentsMargins(16, 10, 16, 10)

        label = QLabel(
            f"<b>Đánh giá & khuyến nghị:</b> "
            f"<span style='font-weight:700'>{text}</span>"
        )

        label.setWordWrap(True)

        label.setStyleSheet(f"""
            QLabel {{
                color: {color};
                font-size: 15px;
                border: none;
                background: transparent;
            }}
        """)

        layout.addWidget(label)

        return box

    def stop_input_preview_video(self):
        if self.input_video_timer.isActive():
            self.input_video_timer.stop()

        if self.input_video_cap is not None:
            try:
                self.input_video_cap.release()
            except Exception:
                pass
            self.input_video_cap = None
            
    def update_input_video_frame(self):
        if self.input_video_cap is None:
            return

        ret, frame = self.input_video_cap.read()

        if not ret:
            self.input_video_cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
            ret, frame = self.input_video_cap.read()

        if ret:
            pix = cv2_to_qpixmap(frame).scaled(
                520, 320, Qt.KeepAspectRatio, Qt.SmoothTransformation
            )
            self.input_preview_label.setPixmap(pix)
            self.input_preview_label.setText("")
    def preview_first_file(self):
        self.stop_input_preview_video()

        if not self.selected_files:
            self.input_preview_label.show()
            self.input_preview_label.setPixmap(QPixmap())
            self.input_preview_label.setText("Chưa có file nào được chọn")
            return

        first_file = self.selected_files[0]
        lower_path = first_file.lower()

        if lower_path.endswith(IMAGE_EXTS):
            img = cv2.imread(first_file)
            if img is not None:
                pix = cv2_to_qpixmap(img).scaled(
                    520, 320, Qt.KeepAspectRatio, Qt.SmoothTransformation
                )
                self.input_preview_label.setPixmap(pix)
                self.input_preview_label.setText("")
            else:
                self.input_preview_label.setPixmap(QPixmap())
                self.input_preview_label.setText("Không thể đọc ảnh preview.")

        elif lower_path.endswith(VIDEO_EXTS):
            self.input_video_cap = cv2.VideoCapture(first_file)

            if not self.input_video_cap.isOpened():
                self.input_preview_label.setPixmap(QPixmap())
                self.input_preview_label.setText("Không thể mở video preview.")
                return

            fps = self.input_video_cap.get(cv2.CAP_PROP_FPS)
            if fps <= 0 or fps > 60:
                fps = 25

            interval = int(1000 / fps)
            self.input_video_timer.start(interval)
            self.update_input_video_frame()

        else:
            self.input_preview_label.setPixmap(QPixmap())
            self.input_preview_label.setText("Định dạng file không hỗ trợ preview.")
            
    def handle_input_media_status(self, status):
        if self.input_preview_player and status == QMediaPlayer.EndOfMedia:
            self.input_preview_player.setPosition(0)
            self.input_preview_player.play()
               
   
    def load_citizen_reports(self):
        self.report_list_widget.clear()
        self.report_data_list = []
        self.current_report_id = None
        self.current_report_status = None
        self.selected_report_data = None
        if hasattr(self, "btn_view_report_detail"):
            self.btn_view_report_detail.setEnabled(False)

        if not os.path.exists(DATABASE_PATH):
            self.report_list_widget.addItem("Chưa có database báo cáo.")
            return

        try:
            self.ensure_report_status_column()

            conn = sqlite3.connect(DATABASE_PATH)
            cursor = conn.cursor()

            query = """
                SELECT 
                    r.id,
                    r.address,
                    r.latitude,
                    r.longitude,
                    r.image_count,
                    r.status,
                    r.created_at,
                    COALESCE(i.image_path, '') AS image_path,
                    COALESCE(i.detected_image_path, '') AS detected_image_path,
                    COALESCE(i.image_name, '') AS image_name
                FROM pothole_reports r
                LEFT JOIN (
                    SELECT report_id, MIN(id) AS first_image_id
                    FROM pothole_report_images
                    GROUP BY report_id
                ) first_i ON r.id = first_i.report_id
                LEFT JOIN pothole_report_images i ON i.id = first_i.first_image_id
            """

            params = []

            if self.report_filter in REPORT_STATUS_INFO:
                query += " WHERE COALESCE(r.status, 'pending') = ?"
                params.append(self.report_filter)

            query += " ORDER BY datetime(r.created_at) DESC, r.id DESC"

            cursor.execute(query, params)
            rows = cursor.fetchall()
            conn.close()

            if not rows:
                self.current_report_id = None
                self.current_report_status = None
                self.selected_report_data = None
                self.btn_view_report_detail.setEnabled(False)
                self.report_list_widget.addItem("Chưa có báo cáo phù hợp.")
                self.draw_reports_on_map()
                return

            for row in rows:
                report_id, address, lat, lng, image_count, status, created_at, image_path, detected_image_path, image_name = row

                status = self.normalize_report_status(status)
                status_text = self.get_report_status_label(status)
                road_name = self.clean_report_road_name(address)
                source_name = self.extract_source_from_report_address(address, image_path or image_name)
                video_time = self.extract_video_time_from_report_address(address)
                time_line = f" | Thời điểm video: {video_time}" if self.is_video_source(source_name) and video_time else ""

                item_text = (
                    f"[{status_text}] #{report_id} | {created_at}\n"
                    f"Tuyến đường: {road_name}\n"
                    f"Nguồn: {source_name}{time_line}\n"
                    f"Tọa độ: {lat}, {lng} | {image_count} ảnh"
                )

                item = QListWidgetItem(item_text)

                report_data = {
                    "id": report_id,
                    "address": address,
                    "road_name": road_name,
                    "source_name": source_name,
                    "video_time": video_time,
                    "image_path": image_path,
                    "detected_image_path": detected_image_path,
                    "image_name": image_name,
                    "latitude": lat,
                    "longitude": lng,
                    "image_count": image_count,
                    "status": status,
                    "created_at": created_at
                }

                item.setData(Qt.UserRole, report_data)
                self.report_data_list.append(report_data)

                item.setForeground(QColor(self.get_report_status_info(status)["text_color"]))

                self.report_list_widget.addItem(item)

            self.draw_reports_on_map()

        except Exception as e:
            self.report_list_widget.addItem(f"Lỗi tải báo cáo: {e}")
       
    def open_report_detail_window(self):
        if not self.selected_report_data:
            QMessageBox.warning(
                self,
                "Chưa chọn báo cáo",
                "Vui lòng chọn một báo cáo trước."
            )
            return

        from ui.report_detail_window import ReportDetailWindow

        self.report_detail_window = ReportDetailWindow(
            self.selected_report_data,
            self
        )
        self.report_detail_window.show()
         
    def on_citizen_report_clicked(self, item):
        report = item.data(Qt.UserRole)

        if not report:
            return

        self.selected_report_data = report
        self.btn_view_report_detail.setEnabled(True)

        self.current_report_id = report["id"]
        self.current_report_status = self.normalize_report_status(report.get("status"))
        report["status"] = self.current_report_status

        status_text = self.get_report_status_label(self.current_report_status)
        self.focus_report_on_map(report)

        self.set_log(
            f"Đã chọn báo cáo #{report['id']}.\n"
            f"Trạng thái: {status_text}."
        )
    
    def on_map_loaded(self, ok):
        self.map_loaded = ok

        if ok:
            self.draw_reports_on_map()
        else:
            self.set_log("Không thể tải bản đồ.")


    def select_report_from_map(self, report_id):
        """Khi admin bấm marker trên map, tự chọn đúng báo cáo ở danh sách bên trái."""
        report_id = str(report_id)

        for index in range(self.report_list_widget.count()):
            item = self.report_list_widget.item(index)
            report = item.data(Qt.UserRole)

            if not report:
                continue

            if str(report.get("id")) == report_id:
                self.report_list_widget.setCurrentItem(item)
                self.report_list_widget.scrollToItem(item)
                self.on_citizen_report_clicked(item)
                return

        self.set_log(f"Không tìm thấy báo cáo #{report_id} trong bộ lọc hiện tại.")

    def draw_reports_on_map(self):
        if not getattr(self, "map_loaded", False):
            return

        reports = []

        for report in getattr(self, "report_data_list", []):
            lat = report.get("latitude")
            lng = report.get("longitude")

            if lat is None or lng is None:
                continue

            try:
                lat = float(lat)
                lng = float(lng)
            except Exception:
                continue

            popup_report = dict(report)
            popup_report["latitude"] = lat
            popup_report["longitude"] = lng

            reports.append({
                "id": report.get("id"),
                "address": report.get("address") or "--",
                "road_name": report.get("road_name") or self.clean_report_road_name(report.get("address")),
                "source_name": report.get("source_name") or self.extract_source_from_report_address(report.get("address"), report.get("image_path") or report.get("image_name")),
                "video_time": report.get("video_time") or self.extract_video_time_from_report_address(report.get("address")),
                "latitude": lat,
                "longitude": lng,
                "status": self.normalize_report_status(report.get("status")),
                "marker_color": self.get_report_status_marker_color(report.get("status")),
                "image_count": report.get("image_count") or 0,
                "created_at": report.get("created_at") or "--",
                "popup_html": self.build_admin_report_popup_html(popup_report),
            })

        reports_json = json.dumps(reports, ensure_ascii=False)

        js = f"""
            (function() {{
                const reports = {reports_json};

                if (typeof L === 'undefined') {{
                    return;
                }}

                const m = (typeof map !== 'undefined') ? map : window.map;

                if (!m) {{
                    return;
                }}

                if (!window.adminReportLayer) {{
                    window.adminReportLayer = L.layerGroup().addTo(m);
                }}

                window.adminReportLayer.clearLayers();

                const points = [];

                reports.forEach(function(r) {{
                    const color = r.marker_color || "#2563EB";

                    const marker = L.circleMarker([r.latitude, r.longitude], {{
                        radius: 9,
                        color: "#ffffff",
                        weight: 2,
                        fillColor: color,
                        fillOpacity: 0.95
                    }}).bindPopup(r.popup_html);

                    marker.on("click", function() {{
                        function callSelect() {{
                            if (window.backend && window.backend.selectReport) {{
                                window.backend.selectReport(String(r.id));
                                return true;
                            }}
                            return false;
                        }}

                        if (callSelect()) return;

                        if (typeof QWebChannel !== "undefined" && window.qt && window.qt.webChannelTransport) {{
                            new QWebChannel(window.qt.webChannelTransport, function(channel) {{
                                window.backend = channel.objects.backend;
                                callSelect();
                            }});
                        }}
                    }});

                    marker.addTo(window.adminReportLayer);
                    points.push([r.latitude, r.longitude]);
                }});

                if (points.length > 0) {{
                    m.fitBounds(points, {{ padding: [40, 40] }});
                }}
            }})();
        """

        self.web_view.page().runJavaScript(js)


    def focus_report_on_map(self, report):
        if not getattr(self, "map_loaded", False):
            return

        lat = report.get("latitude")
        lng = report.get("longitude")

        if lat is None or lng is None:
            return

        try:
            lat = float(lat)
            lng = float(lng)
        except Exception:
            return

        report_id = report.get("id")
        popup_report = dict(report)
        popup_report["latitude"] = lat
        popup_report["longitude"] = lng
        popup = self.build_admin_report_popup_html(popup_report)

        popup_json = json.dumps(popup, ensure_ascii=False)

        js = f"""
            (function() {{
                if (typeof L === 'undefined') {{
                    return;
                }}

                const m = (typeof map !== 'undefined') ? map : window.map;

                if (!m) {{
                    return;
                }}

                if (!window.adminSelectedLayer) {{
                    window.adminSelectedLayer = L.layerGroup().addTo(m);
                }}

                window.adminSelectedLayer.clearLayers();

                const marker = L.marker([{lat}, {lng}]).bindPopup({popup_json});
                marker.addTo(window.adminSelectedLayer);

                m.setView([{lat}, {lng}], 18);
                marker.openPopup();
            }})();
        """

        self.web_view.page().runJavaScript(js)


    def search_location_by_button(self):
        keyword = self.txt_search.text().strip()

        if not keyword:
            QMessageBox.warning(self, "Thiếu thông tin", "Vui lòng nhập tên đường hoặc địa chỉ.")
            return

        keyword_json = json.dumps(keyword, ensure_ascii=False)

        js = f"""
            (function() {{
                const keyword = {keyword_json};

                if (typeof L === 'undefined') {{
                    return;
                }}

                const m = (typeof map !== 'undefined') ? map : window.map;

                if (!m) {{
                    return;
                }}

                fetch("https://nominatim.openstreetmap.org/search?format=json&q=" + encodeURIComponent(keyword))
                    .then(response => response.json())
                    .then(data => {{
                        if (!data || data.length === 0) {{
                            alert("Không tìm thấy địa chỉ.");
                            return;
                        }}

                        const result = data[0];
                        const lat = parseFloat(result.lat);
                        const lon = parseFloat(result.lon);

                        m.setView([lat, lon], 17);

                        if (!window.adminSearchLayer) {{
                            window.adminSearchLayer = L.layerGroup().addTo(m);
                        }}

                        window.adminSearchLayer.clearLayers();

                        L.marker([lat, lon])
                            .addTo(window.adminSearchLayer)
                            .bindPopup(result.display_name)
                            .openPopup();
                    }})
                    .catch(error => {{
                        alert("Không thể tìm kiếm địa chỉ.");
                    }});
            }})();
        """

        self.web_view.page().runJavaScript(js)
    
    def stop_all_input_videos(self):
        for timer in self.input_video_timers.values():
            try:
                timer.stop()
            except Exception:
                pass

        for cap in self.input_video_caps.values():
            try:
                cap.release()
            except Exception:
                pass

        self.input_video_timers = {}
        self.input_video_caps = {}
        self.input_video_labels = {}
        
    def ensure_report_status_column(self):
        if not os.path.exists(DATABASE_PATH):
            return

        conn = sqlite3.connect(DATABASE_PATH)
        cursor = conn.cursor()

        cursor.execute("PRAGMA table_info(pothole_reports)")
        columns = [column[1] for column in cursor.fetchall()]

        if "status" not in columns:
            cursor.execute("""
                ALTER TABLE pothole_reports
                ADD COLUMN status TEXT DEFAULT 'pending'
            """)

        # Đảm bảo database cũ vẫn có bảng lưu từng ảnh/khung hình báo cáo.
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS pothole_report_images (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                report_id INTEGER NOT NULL,
                image_path TEXT,
                image_name TEXT,
                detected_image_path TEXT,
                analysis_html TEXT,
                area_m2 REAL DEFAULT 0,
                setup_name TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (report_id) REFERENCES pothole_reports(id)
            )
        """)

        cursor.execute("PRAGMA table_info(pothole_report_images)")
        image_columns = [column[1] for column in cursor.fetchall()]

        if "detected_image_path" not in image_columns:
            cursor.execute("ALTER TABLE pothole_report_images ADD COLUMN detected_image_path TEXT")

        if "analysis_html" not in image_columns:
            cursor.execute("ALTER TABLE pothole_report_images ADD COLUMN analysis_html TEXT")

        if "area_m2" not in image_columns:
            cursor.execute("ALTER TABLE pothole_report_images ADD COLUMN area_m2 REAL DEFAULT 0")

        if "setup_name" not in image_columns:
            cursor.execute("ALTER TABLE pothole_report_images ADD COLUMN setup_name TEXT")

        conn.commit()
        conn.close()


    def set_report_filter(self, filter_value):
        self.report_filter = filter_value
        self.update_report_filter_button_styles()
        self.load_citizen_reports()


    def update_report_filter_button_styles(self):
        buttons = getattr(self, "status_filter_buttons", {})
        if not buttons:
            return

        for filter_value, button in buttons.items():
            is_active = self.report_filter == filter_value

            if filter_value == "all":
                if is_active:
                    button.setStyleSheet("""
                        QPushButton {
                            background-color: #2563EB;
                            color: white;
                            border: 1px solid #2563EB;
                            border-radius: 8px;
                            padding: 7px 10px;
                            font-size: 12px;
                            font-weight: 900;
                        }
                    """)
                else:
                    button.setStyleSheet("""
                        QPushButton {
                            background-color: #FFFFFF;
                            color: #111827;
                            border: 1px solid #CBD5E1;
                            border-radius: 8px;
                            padding: 7px 10px;
                            font-size: 12px;
                            font-weight: 800;
                        }
                        QPushButton:hover { background-color: #F8FAFC; }
                    """)
                continue

            info = REPORT_STATUS_INFO[filter_value]
            if is_active:
                button.setStyleSheet(f"""
                    QPushButton {{
                        background-color: {info['marker_color']};
                        color: white;
                        border: 1px solid {info['marker_color']};
                        border-radius: 8px;
                        padding: 7px 10px;
                        font-size: 12px;
                        font-weight: 900;
                    }}
                """)
            else:
                button.setStyleSheet(f"""
                    QPushButton {{
                        background-color: {info['bg_color']};
                        color: {info['text_color']};
                        border: 1px solid {info['border_color']};
                        border-radius: 8px;
                        padding: 7px 10px;
                        font-size: 12px;
                        font-weight: 800;
                    }}
                    QPushButton:hover {{ background-color: #FFFFFF; }}
                """)

    # =========================
    # EXPORT REPORTS PDF / EXCEL
    # =========================
    def get_selected_export_filter(self):
        if not hasattr(self, "cmb_export_status"):
            return self.report_filter

        selected = self.cmb_export_status.currentData()
        if selected == "current":
            return self.report_filter
        return selected or "all"

    def get_export_filter_label(self, status_filter):
        if status_filter in REPORT_STATUS_INFO:
            return REPORT_STATUS_INFO[status_filter]["label"]
        return "Tất cả"

    def normalize_html_to_text(self, value):
        text = str(value or "")
        text = re.sub(r"<\s*br\s*/?\s*>", "\n", text, flags=re.IGNORECASE)
        text = re.sub(r"</\s*p\s*>", "\n", text, flags=re.IGNORECASE)
        text = re.sub(r"</\s*div\s*>", "\n", text, flags=re.IGNORECASE)
        text = re.sub(r"<[^>]+>", "", text)
        text = html.unescape(text)
        lines = [line.strip() for line in text.splitlines()]
        return "\n".join(line for line in lines if line)

    def format_frame_time_for_export(self, frame_time):
        if frame_time in (None, ""):
            return ""
        try:
            total_seconds = float(frame_time)
        except Exception:
            return str(frame_time)

        minutes = int(total_seconds // 60)
        seconds = total_seconds - minutes * 60
        hours = minutes // 60
        minutes = minutes % 60

        if hours > 0:
            return f"{hours:02d}:{minutes:02d}:{seconds:05.2f}"
        return f"{minutes:02d}:{seconds:05.2f}"

    def get_export_reports_data(self, status_filter=None):
        """Lấy dữ liệu báo cáo để xuất Excel/PDF, có thể lọc theo trạng thái."""
        status_filter = status_filter or self.get_selected_export_filter()
        self.ensure_report_status_column()

        conn = sqlite3.connect(DATABASE_PATH)
        cursor = conn.cursor()

        cursor.execute("PRAGMA table_info(pothole_reports)")
        report_columns = {column[1] for column in cursor.fetchall()}

        cursor.execute("PRAGMA table_info(pothole_report_images)")
        image_columns = {column[1] for column in cursor.fetchall()}

        def report_expr(column_name, default="''"):
            return f"COALESCE(r.{column_name}, {default})" if column_name in report_columns else default

        def image_expr(column_name, default="''"):
            return f"COALESCE(i.{column_name}, {default})" if column_name in image_columns else default

        query = f"""
            SELECT
                r.id,
                COALESCE(r.address, '') AS address,
                r.latitude,
                r.longitude,
                COALESCE(r.image_count, 0) AS image_count,
                COALESCE(r.status, 'pending') AS status,
                COALESCE(r.created_at, '') AS created_at,
                {report_expr('user_id', 'NULL')} AS user_id,
                {report_expr('reporter_user_id', 'NULL')} AS reporter_user_id,
                {report_expr('reviewed_by', 'NULL')} AS reviewed_by,
                {report_expr('reviewed_at')} AS reviewed_at,
                {report_expr('analysis_html')} AS report_analysis_html,
                {report_expr('video_name')} AS report_video_name,
                {report_expr('frame_time', 'NULL')} AS report_frame_time,
                {report_expr('confidence', 'NULL')} AS confidence,
                {report_expr('gps_source')} AS gps_source,
                {report_expr('report_type')} AS report_type,
                {image_expr('image_path')} AS image_path,
                {image_expr('image_name')} AS image_name,
                {image_expr('detected_image_path')} AS detected_image_path,
                {image_expr('analysis_html')} AS image_analysis_html,
                {image_expr('area_m2', '0')} AS area_m2,
                {image_expr('setup_name')} AS setup_name
            FROM pothole_reports r
            LEFT JOIN (
                SELECT report_id, MIN(id) AS first_image_id
                FROM pothole_report_images
                GROUP BY report_id
            ) first_i ON r.id = first_i.report_id
            LEFT JOIN pothole_report_images i ON i.id = first_i.first_image_id
        """

        params = []
        if status_filter in REPORT_STATUS_INFO:
            query += " WHERE COALESCE(r.status, 'pending') = ?"
            params.append(status_filter)

        query += " ORDER BY datetime(r.created_at) DESC, r.id DESC"

        cursor.execute(query, params)
        db_rows = cursor.fetchall()
        conn.close()

        headers = [
            "Mã báo cáo",
            "Trạng thái",
            "Tuyến đường",
            "Nguồn",
            "Thời điểm video",
            "Vĩ độ",
            "Kinh độ",
            "Số file/ảnh",
            "Thời gian gửi",
            "Người gửi",
            "Người xử lý",
            "Thời gian cập nhật",
            "Độ tin cậy",
            "Nguồn GPS",
            "Loại báo cáo",
            "Diện tích ước lượng (m²)",
            "Setup camera",
            "Ảnh gốc",
            "Ảnh phát hiện",
            "Thông tin phân tích",
        ]

        rows = []
        for row in db_rows:
            (
                report_id,
                address,
                latitude,
                longitude,
                image_count,
                status,
                created_at,
                user_id,
                reporter_user_id,
                reviewed_by,
                reviewed_at,
                report_analysis_html,
                report_video_name,
                report_frame_time,
                confidence,
                gps_source,
                report_type,
                image_path,
                image_name,
                detected_image_path,
                image_analysis_html,
                area_m2,
                setup_name,
            ) = row

            status = self.normalize_report_status(status)
            road_name = self.clean_report_road_name(address)

            source_name = self.extract_source_from_report_address(address, image_path or image_name or report_video_name)
            if (not source_name or source_name == "--") and report_video_name:
                source_name = self.get_media_source_name(report_video_name)

            video_time = self.extract_video_time_from_report_address(address)
            if not video_time and self.is_video_source(source_name):
                video_time = self.format_frame_time_for_export(report_frame_time)

            analysis_html = image_analysis_html or report_analysis_html or ""
            analysis_text = self.normalize_html_to_text(analysis_html)

            try:
                area_value = float(area_m2 or 0)
            except Exception:
                area_value = 0.0

            try:
                confidence_value = "" if confidence in (None, "") else float(confidence)
            except Exception:
                confidence_value = confidence

            sender = reporter_user_id if reporter_user_id not in (None, "") else user_id

            rows.append([
                report_id,
                self.get_report_status_label(status),
                road_name,
                source_name,
                video_time if self.is_video_source(source_name) else "",
                latitude,
                longitude,
                image_count,
                created_at,
                sender if sender not in (None, "") else "",
                reviewed_by if reviewed_by not in (None, "") else "",
                reviewed_at or "",
                confidence_value,
                gps_source or "",
                report_type or "",
                area_value,
                setup_name or "",
                image_path or "",
                detected_image_path or "",
                analysis_text,
            ])

        return headers, rows

    def build_export_summary_html(self, headers, rows, status_filter):
        generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        status_label = self.get_export_filter_label(status_filter)

        table_header = "".join(
            f"<th>{html.escape(header)}</th>"
            for header in headers
        )

        body_rows = []
        for row in rows:
            cells = []
            for value in row:
                text = html.escape(str(value if value is not None else "")).replace("\n", "<br>")
                cells.append(f"<td>{text}</td>")
            body_rows.append("<tr>" + "".join(cells) + "</tr>")

        if not body_rows:
            body_rows.append(
                f"<tr><td colspan='{len(headers)}' style='text-align:center;padding:18px'>Không có dữ liệu phù hợp.</td></tr>"
            )

        return f"""
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-family: 'Segoe UI', Arial, sans-serif; color: #111827; }}
                h1 {{ font-size: 20px; color: #1F2937; margin-bottom: 4px; }}
                .subtitle {{ font-size: 11px; color: #475569; margin-bottom: 12px; }}
                table {{ border-collapse: collapse; width: 100%; font-size: 8px; }}
                th {{ background: #2563EB; color: white; padding: 5px; border: 1px solid #CBD5E1; }}
                td {{ padding: 4px; border: 1px solid #CBD5E1; vertical-align: top; }}
                tr:nth-child(even) {{ background: #F8FAFC; }}
            </style>
        </head>
        <body>
            <h1>THỐNG KÊ DANH SÁCH BÁO CÁO Ổ GÀ</h1>
            <div class="subtitle">
                Trạng thái xuất: <b>{html.escape(status_label)}</b> | Số báo cáo: <b>{len(rows)}</b> | Thời gian xuất: {generated_at}
            </div>
            <table>
                <thead><tr>{table_header}</tr></thead>
                <tbody>{''.join(body_rows)}</tbody>
            </table>
        </body>
        </html>
        """

    def make_export_default_filename(self, extension):
        status_filter = self.get_selected_export_filter()
        status_name = "tat_ca" if status_filter not in REPORT_STATUS_INFO else status_filter
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return os.path.join(
            os.path.expanduser("~"),
            f"thong_ke_bao_cao_o_ga_{status_name}_{timestamp}.{extension}"
        )

    def export_reports_to_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except Exception:
            QMessageBox.warning(
                self,
                "Thiếu thư viện Excel",
                "Chưa cài thư viện openpyxl.\nHãy chạy lệnh: pip install openpyxl"
            )
            return

        status_filter = self.get_selected_export_filter()
        headers, rows = self.get_export_reports_data(status_filter)

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Xuất thống kê Excel",
            self.make_export_default_filename("xlsx"),
            "Excel Workbook (*.xlsx)"
        )

        if not file_path:
            return
        if not file_path.lower().endswith(".xlsx"):
            file_path += ".xlsx"

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Thong ke o ga"

            status_label = self.get_export_filter_label(status_filter)
            generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            ws.cell(row=1, column=1).value = "THỐNG KÊ DANH SÁCH BÁO CÁO Ổ GÀ"
            ws.cell(row=1, column=1).font = Font(size=16, bold=True, color="1F2937")
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
            ws.cell(row=2, column=1).value = f"Trạng thái xuất: {status_label} | Số báo cáo: {len(rows)} | Thời gian xuất: {generated_at}"
            ws.cell(row=2, column=1).font = Font(size=11, italic=True, color="475569")
            ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

            header_row = 4
            for col_index, header in enumerate(headers, start=1):
                cell = ws.cell(row=header_row, column=col_index)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2563EB")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            thin = Side(style="thin", color="CBD5E1")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for row_index, row_values in enumerate(rows, start=header_row + 1):
                for col_index, value in enumerate(row_values, start=1):
                    cell = ws.cell(row=row_index, column=col_index)
                    cell.value = value
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = border

            for col_index in range(1, len(headers) + 1):
                header = headers[col_index - 1]
                letter = get_column_letter(col_index)
                if header in ("Thông tin phân tích", "Ảnh gốc", "Ảnh phát hiện", "Tuyến đường"):
                    width = 36
                elif header in ("Thời gian gửi", "Thời gian cập nhật"):
                    width = 20
                elif header in ("Vĩ độ", "Kinh độ"):
                    width = 14
                else:
                    width = 16
                ws.column_dimensions[letter].width = width

            ws.row_dimensions[1].height = 24
            ws.row_dimensions[2].height = 20
            ws.row_dimensions[4].height = 34
            ws.freeze_panes = "A5"
            ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{max(4, len(rows) + 4)}"

            wb.save(file_path)

            QMessageBox.information(
                self,
                "Xuất Excel thành công",
                f"Đã xuất {len(rows)} báo cáo ra file:\n{file_path}"
            )
            self.set_log(f"Đã xuất Excel: {file_path}")

        except Exception as error:
            QMessageBox.critical(
                self,
                "Lỗi xuất Excel",
                f"Không thể xuất file Excel.\n\nChi tiết: {error}"
            )

    def export_reports_to_pdf(self):
        status_filter = self.get_selected_export_filter()
        headers, rows = self.get_export_reports_data(status_filter)

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Xuất thống kê PDF",
            self.make_export_default_filename("pdf"),
            "PDF Files (*.pdf)"
        )

        if not file_path:
            return
        if not file_path.lower().endswith(".pdf"):
            file_path += ".pdf"

        try:
            document = QTextDocument()
            document.setHtml(self.build_export_summary_html(headers, rows, status_filter))

            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFormat(QPrinter.PdfFormat)
            printer.setOutputFileName(file_path)
            printer.setPageSize(QPrinter.A4)
            printer.setOrientation(QPrinter.Landscape)
            printer.setPageMargins(8, 8, 8, 8, QPrinter.Millimeter)

            document.print_(printer)

            QMessageBox.information(
                self,
                "Xuất PDF thành công",
                f"Đã xuất {len(rows)} báo cáo ra file:\n{file_path}"
            )
            self.set_log(f"Đã xuất PDF: {file_path}")

        except Exception as error:
            QMessageBox.critical(
                self,
                "Lỗi xuất PDF",
                f"Không thể xuất file PDF.\n\nChi tiết: {error}"
            )

    def approve_current_report(self):
        self.update_current_report_status("approved")

    def update_input_video_card_frame(self, file_path):
        cap = self.input_video_caps.get(file_path)
        label = self.input_video_labels.get(file_path)

        if cap is None or label is None:
            return

        try:
            if label.parent() is None:
                return

            ret, frame = cap.read()

            if not ret:
                cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
                ret, frame = cap.read()

            if ret:
                pix = cv2_to_qpixmap(frame).scaled(
                    700, 420, Qt.KeepAspectRatio, Qt.SmoothTransformation
                )
                label.setPixmap(pix)

        except RuntimeError:
            timer = self.input_video_timers.get(file_path)
            if timer:
                timer.stop()
            
    def play_input_video(self, video_path, speed=1.0):
        cap = self.input_video_caps.get(video_path)
        timer = self.input_video_timers.get(video_path)

        if cap is None or timer is None:
            return

        fps = cap.get(cv2.CAP_PROP_FPS)
        if fps <= 0 or fps > 60:
            fps = 25

        interval = int(1000 / (fps * speed))
        timer.start(max(interval, 1))


    def pause_input_video(self, video_path):
        timer = self.input_video_timers.get(video_path)
        if timer:
            timer.stop()


    def stop_input_video(self, video_path):
        cap = self.input_video_caps.get(video_path)
        timer = self.input_video_timers.get(video_path)

        if timer:
            timer.stop()

        if cap:
            cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
            self.update_input_video_card_frame(video_path)

    def on_video_progress(self, video_path, value):
        progress_bar = self.video_progress_bars.get(video_path)
        if progress_bar:
            progress_bar.setValue(value)
        
    def on_video_error(self, video_path, err):
        status_label = self.video_status_labels.get(video_path)
        progress_bar = self.video_progress_bars.get(video_path)

        if status_label:
            status_label.setText(f"Lỗi khi xử lý video: {err}")
            status_label.setStyleSheet("""
                QLabel {
                    color: #DC2626;
                    font-size: 12px;
                    font-weight: 700;
                }
            """)

        if progress_bar:
            progress_bar.setValue(0)
